"""Excel-Import (openpyxl) und Erzeugung der Beispiel-Vorlage."""
from __future__ import annotations

import datetime
import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

MAX_ROWS = 5000
MAX_DISTINCT = 80


def _cell_to_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return value.isoformat(sep=" ") if isinstance(value, datetime.datetime) else value.isoformat()
    return str(value).strip()


def _dedupe_headers(raw: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    out = []
    for i, h in enumerate(raw):
        name = h.strip() or f"Spalte_{i + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name} ({seen[name]})"
        else:
            seen[name] = 1
        out.append(name)
    return out


def parse_workbook(data: bytes, sheet_name: str | None = None) -> dict:
    """Parst ein XLSX; erste nicht-leere Zeile = Kopfzeile.

    Rückgabe: {sheets, sheet, headers, rows, row_count, distinct, truncated}
    Jede Zeile enthält zusätzlich "_row" (Original-Zeilennummer in Excel).
    """
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    sheets = wb.sheetnames
    if not sheets:
        raise ValueError("Die Arbeitsmappe enthält keine Tabellenblätter.")
    active = sheet_name if sheet_name in sheets else sheets[0]
    ws = wb[active]

    headers: list[str] = []
    rows: list[dict] = []
    truncated = False
    header_row_no = 0

    for row_no, row in enumerate(ws.iter_rows(values_only=True), start=1):
        vals = [_cell_to_str(c) for c in (row or ())]
        if not headers:
            if any(vals):
                headers = _dedupe_headers(vals)
                header_row_no = row_no
            continue
        if not any(vals):
            continue
        entry = {"_row": row_no}
        for i, h in enumerate(headers):
            entry[h] = vals[i] if i < len(vals) else ""
        rows.append(entry)
        if len(rows) >= MAX_ROWS:
            truncated = True
            break

    wb.close()
    if not headers:
        raise ValueError(f"Keine Kopfzeile im Blatt '{active}' gefunden.")

    distinct: dict[str, list[str]] = {}
    for h in headers:
        vals = sorted({r[h] for r in rows if r.get(h)})
        if 0 < len(vals) <= MAX_DISTINCT:
            distinct[h] = vals

    return {
        "sheets": sheets,
        "sheet": active,
        "header_row": header_row_no,
        "headers": headers,
        "rows": rows,
        "row_count": len(rows),
        "distinct": distinct,
        "truncated": truncated,
    }


_HEAD_FONT = Font(bold=True, color="FFFFFF", name="Bahnschrift")
_HEAD_FILL = PatternFill("solid", fgColor="1D2226")


def _style_sheet(ws, widths: list[int]) -> None:
    for col, width in enumerate(widths, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEAD_FONT
        cell.fill = _HEAD_FILL
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"


def template_workbook(templates: list[dict]) -> bytes:
    """Import-Vorlage passend zu konkreten FortiPAM-Templates.

    Spalten: Basisfelder + Vereinigungsmenge der Template-Felder (Felder vom
    Typ target-address entfallen — die Adresse kommt aus der Target-Spalte).
    Als Secret-Typ wird der exakte Template-Name vorgegeben, damit die
    Zuordnung im Mapping automatisch greift.
    """
    base = ["Name", "Adresse", "Secret-Typ", "Ordner", "Beschreibung"]
    field_cols: list[str] = []
    for tpl in templates:
        for f in tpl.get("field", []) or []:
            name = str(f.get("name", "")).strip()
            if not name or f.get("type") == "target-address":
                continue
            if name not in field_cols:
                field_cols.append(name)

    wb = Workbook()
    ws = wb.active
    ws.title = "Import"
    headers = base + field_cols
    ws.append(headers)
    _style_sheet(ws, [26, 22, 30, 24, 30] + [20] * len(field_cols))

    for i, tpl in enumerate(templates, start=1):
        row = {"Name": f"beispiel-{i:02d}", "Adresse": f"10.0.0.{i}",
               "Secret-Typ": tpl.get("name", ""), "Ordner": "Import/Beispiel",
               "Beschreibung": f"Beispielzeile für {tpl.get('name', '')}"}
        for f in tpl.get("field", []) or []:
            name = str(f.get("name", "")).strip()
            if name in field_cols:
                row[name] = ""
        ws.append([row.get(h, "") for h in headers])
    return _save(wb)


def inventory_workbook(inv: dict) -> bytes:
    """Exportiert das Inventar als Arbeitsmappe mit einem Blatt je Objekttyp."""
    wb = Workbook()
    wb.remove(wb.active)

    def sheet(title, headers, widths, rows):
        ws = wb.create_sheet(title)
        ws.append(headers)
        _style_sheet(ws, widths)
        for r in rows:
            ws.append(r)

    sheet("Targets",
          ["Name", "Adresse", "Template", "Klassifizierung", "Domäne", "Beschreibung"],
          [28, 22, 30, 22, 22, 34],
          [[t.get("name", ""), t.get("address", ""), t.get("template", ""),
            t.get("class", ""), t.get("domain", ""), t.get("description", "")]
           for t in inv.get("targets", [])])
    sheet("Secrets",
          ["ID", "Name", "Ordner", "Template", "Target", "Beschreibung"],
          [8, 28, 28, 30, 28, 34],
          [[s.get("id", ""), s.get("name", ""), s.get("folder_path", ""),
            s.get("template", ""), s.get("target", ""), s.get("description", "")]
           for s in inv.get("secrets", [])])
    sheet("Ordner",
          ["ID", "Pfad", "Übergeordnet (ID)"],
          [8, 40, 18],
          [[f.get("id", ""), f.get("path", ""), f.get("parent-folder", "")]
           for f in inv.get("folders", [])])
    sheet("Templates",
          ["Name", "Server-Info", "Felder"],
          [32, 16, 60],
          [[t.get("name", ""), t.get("server-info", ""),
            ", ".join(f"{f.get('name')} ({f.get('type')})" for f in (t.get("field") or []))]
           for t in inv.get("templates", [])])
    sheet("Klassifizierungen",
          ["Name", "Beschreibung"],
          [28, 50],
          [[t.get("name", ""), t.get("description", "")]
           for t in inv.get("class_tags", [])])
    return _save(wb)


def _save(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def sample_workbook() -> bytes:
    """Erzeugt eine Beispiel-Importvorlage als XLSX."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Import"

    headers = ["Name", "Adresse", "Secret-Typ", "Benutzername", "Passwort",
               "Domäne", "Ordner", "Beschreibung"]
    widths = [26, 22, 24, 20, 20, 20, 24, 34]
    ws.append(headers)

    head_font = Font(bold=True, color="FFFFFF", name="Bahnschrift")
    head_fill = PatternFill("solid", fgColor="1D2226")
    for col, width in enumerate(widths, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = head_font
        cell.fill = head_fill
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"

    ws.append(["srv-linux-01", "10.10.1.21", "linux", "root", "GeheimesPasswort1!",
               "", "Linux/Produktion", "Webserver Frontend"])
    ws.append(["srv-win-01", "10.10.2.15", "windows", "Administrator", "GeheimesPasswort2!",
               "corp.example.com", "Windows/Produktion", "Domain Controller"])
    ws.append(["sw-core-01", "10.10.0.2", "netzwerk", "admin", "GeheimesPasswort3!",
               "", "Netzwerk", "Core Switch"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
